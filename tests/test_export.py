"""Excel shaping and the export gate, on synthetic frames and secrets files."""

from io import BytesIO

import pandas as pd
import pyarrow as pa
import pytest
from openpyxl import load_workbook
from openpyxl.cell.cell import ERROR_CODES
from streamlit.testing.v1 import AppTest

from edscrib.auth import MESSAGE_REJECTED, MESSAGE_UNAVAILABLE
from edscrib.export import _MAX_COLUMN_WIDTH, build_workbook

USERS = {"annotator_a": "sept-chevaux"}
SHEET = "2023-01_avc_annotator_a"
LABEL = "**TÉLÉCHARGER LES DONNÉES ANNOTÉES**"
NOTICE = "## IMPORTANT\nWhat leaving with a gold standard commits you to."
LONGEST = "un commentaire nettement plus long que les autres"

FRAME = pd.DataFrame(
    {
        "n": [1, 2, 3],
        "note_estimate_a": ["oui", "non", "oui"],
        "note_comment_a": ["court", LONGEST, "-"],
    }
)


@pytest.fixture
def secrets(tmp_path):
    path = tmp_path / "secrets.toml"
    table = "\n".join(f'{name} = "{word}"' for name, word in USERS.items())
    path.write_text(f"[users]\n{table}\n")
    return path


def sheet_of(data=FRAME, sheet_name=SHEET):
    return load_workbook(build_workbook(data, sheet_name))[sheet_name]


def test_build_workbook_writes_the_frame_under_the_sheet_name():
    read = pd.read_excel(build_workbook(FRAME, SHEET), sheet_name=SHEET)

    assert read.columns.tolist() == FRAME.columns.tolist()
    assert read.to_dict("list") == FRAME.to_dict("list")


def test_build_workbook_leaves_the_index_out():
    sheet = sheet_of()

    assert [cell.value for cell in sheet[1]] == FRAME.columns.tolist()


def test_build_workbook_rewinds_the_buffer():
    assert build_workbook(FRAME, SHEET).tell() == 0


def test_build_workbook_widens_each_column_to_its_longest_value():
    sheet = sheet_of()
    widths = [sheet.column_dimensions[letter].width for letter in ("A", "B", "C")]

    assert widths[2] == pytest.approx((len(LONGEST) + 2) * 1.1)
    assert widths[2] == max(widths)
    assert all(width > 0 for width in widths)


def test_build_workbook_caps_a_column_widened_by_a_long_comment():
    frame = pd.DataFrame({"n": [1], "note_comment_a": ["x" * 400]})
    sheet = load_workbook(build_workbook(frame, SHEET)).worksheets[0]

    assert sheet.column_dimensions["B"].width == _MAX_COLUMN_WIDTH
    assert _MAX_COLUMN_WIDTH <= 255


def test_build_workbook_fills_the_header_row_and_no_other():
    sheet = sheet_of()

    assert sheet["A1"].fill.fgColor.rgb.endswith("E5E5E5")
    assert sheet["A2"].fill.fill_type is None


def test_build_workbook_centres_and_borders_every_cell():
    sheet = sheet_of()

    for coordinate in ("A1", "C3"):
        cell = sheet[coordinate]

        assert cell.alignment.horizontal == "center"
        assert cell.alignment.vertical == "center"
        assert cell.border.left.style == "thin"
        assert cell.border.bottom.color.rgb.endswith("A5A5A5")


def test_build_workbook_filters_over_the_whole_table():
    assert sheet_of().auto_filter.ref == "A1:C4"


def test_build_workbook_refuses_a_sheet_name_openpyxl_would_not_carry():
    name = "x" * 32

    with pytest.raises(ValueError, match="32 characters"):
        build_workbook(FRAME, name)


def test_build_workbook_refuses_a_value_openpyxl_would_truncate():
    frame = pd.DataFrame({"n": [1], "note_comment_a": ["x" * 32768]})

    with pytest.raises(ValueError, match="note_comment_a"):
        build_workbook(frame, SHEET)


def test_build_workbook_refuses_a_truncated_value_an_arrow_backed_column_holds():
    column = pd.Series(["x" * 32768], dtype=pd.ArrowDtype(pa.string()))
    frame = pd.DataFrame({"n": [1], "note_comment_a": column})

    with pytest.raises(ValueError, match="note_comment_a"):
        build_workbook(frame, SHEET)


def test_build_workbook_refuses_a_control_character_by_position_not_by_value():
    comments = ["court", "patient \x0b Mme Dupont"]
    frame = pd.DataFrame({"n": [1, 2], "note_comment_a": comments})

    with pytest.raises(ValueError) as refusal:
        build_workbook(frame, SHEET)

    assert "note_comment_a" in str(refusal.value)
    assert "position 1" in str(refusal.value)
    assert "Dupont" not in str(refusal.value)


def test_build_workbook_refuses_a_control_character_in_a_column_name():
    frame = pd.DataFrame({"n": [1], "note\x0bcomment_a": ["court"]})

    with pytest.raises(ValueError, match="column name"):
        build_workbook(frame, SHEET)


def test_build_workbook_writes_a_comment_opening_on_an_equals_sign_as_text():
    shorthand = ["=> voir le compte-rendu", "= idem", "=/= du CR initial"]
    frame = pd.DataFrame({"n": [1, 2, 3], "note_comment_a": shorthand})

    buffer = build_workbook(frame, SHEET)
    sheet = load_workbook(buffer).worksheets[0]

    assert [cell.data_type for cell in sheet["B"][1:]] == ["s", "s", "s"]

    buffer.seek(0)

    assert pd.read_excel(buffer, sheet_name=SHEET)["note_comment_a"].tolist() == shorthand


@pytest.mark.parametrize("code", ERROR_CODES)
def test_build_workbook_writes_a_comment_equal_to_an_error_code_as_text(code):
    frame = pd.DataFrame({"n": [1], "note_comment_a": [code]})

    cell = load_workbook(build_workbook(frame, SHEET)).worksheets[0]["B2"]

    assert cell.data_type == "s"
    assert cell.value == code


@pytest.mark.parametrize("name", ["sheet", "SHEET", "ShEeT", "Sheet"])
def test_build_workbook_styles_the_sheet_openpyxl_renames_under_it(name):
    sheet = load_workbook(build_workbook(FRAME, name)).worksheets[0]

    assert sheet["A1"].fill.fgColor.rgb.endswith("E5E5E5")
    assert sheet["A1"].border.left.style == "thin"
    assert sheet.freeze_panes == "A2"


def test_build_workbook_exports_a_frame_with_no_rows():
    empty = FRAME.iloc[:0]
    sheet = sheet_of(empty)

    assert [cell.value for cell in sheet[1]] == FRAME.columns.tolist()
    assert sheet.auto_filter.ref == "A1:C1"


def _render(path, data, options):
    from edscrib.export import download_form

    download_form(path, data, **options)


def run_form(secrets, data=FRAME, **options):
    app = AppTest.from_function(
        _render,
        kwargs={
            "path": str(secrets),
            "data": data,
            "options": {
                "filename": SHEET,
                "label": LABEL,
                "info": NOTICE,
                "icon": ":material/download:",
            }
            | options,
        },
    )
    return app.run()


def submit(app, username, password):
    app.text_input[0].input(username)
    return app.text_input[1].input(password).run()


def error_body(app):
    """The message `st.error` renders, minus a leading emoji it turns into an icon."""
    return app.error[0].value


def test_download_form_asks_for_a_pair_and_serves_nothing():
    app = run_form("absent.toml")

    assert not app.exception
    assert [field.label for field in app.text_input] == ["Identifiant", "Mot de passe"]
    assert not app.download_button
    assert not app.error
    assert not app.markdown


def test_download_form_serves_the_export_on_a_matching_pair(secrets):
    app = submit(run_form(secrets), "annotator_a", USERS["annotator_a"])

    assert not app.exception
    assert not app.error

    button = app.download_button[0]

    assert button.label == LABEL
    assert button.proto.url.endswith(".xlsx")
    assert button.proto.ignore_rerun


def test_download_form_holds_the_notice_back_until_the_pair_matches(secrets):
    before = run_form(secrets)

    assert NOTICE not in [block.value for block in before.markdown]

    after = submit(before, "annotator_a", USERS["annotator_a"])

    assert NOTICE in [block.value for block in after.markdown]


def test_download_form_renders_no_notice_when_none_is_passed(secrets):
    app = submit(run_form(secrets, info=""), "annotator_a", USERS["annotator_a"])

    assert app.download_button
    assert not app.markdown


def test_download_form_rejects_a_wrong_password_without_naming_which_field(secrets):
    app = submit(run_form(secrets), "annotator_a", "trois-oranges")
    message = error_body(app)

    assert not app.exception
    assert not app.download_button
    assert "utilisateur" in message.lower()
    assert "mot de passe" in message.lower()
    assert MESSAGE_REJECTED.endswith(message)


def test_download_form_rejects_an_unknown_user_with_the_same_message(secrets):
    unknown = submit(run_form(secrets), "annotator_c", USERS["annotator_a"])
    wrong = submit(run_form(secrets), "annotator_a", "trois-oranges")

    assert not unknown.download_button
    assert error_body(unknown) == error_body(wrong)


def _render_two(path, data, options):
    from edscrib.export import download

    download(path, data, **options, key_button="button-export-a")
    download(path, data, **options, key_button="button-export-b")


def test_download_renders_twice_on_one_page_under_distinct_keys(secrets):
    app = AppTest.from_function(
        _render_two,
        kwargs={
            "path": str(secrets),
            "data": FRAME,
            "options": {"filename": SHEET, "label": LABEL},
        },
    ).run()

    assert not app.exception
    assert len(app.button) == 2


def unreadable(tmp_path, kind):
    path = tmp_path / "secrets.toml"

    if kind == "absent":
        return path

    if kind == "directory":
        path.mkdir()
    elif kind == "encoding":
        path.write_bytes('[users]\nannotator_a = "cl\xe9"\n'.encode("latin-1"))
    elif kind == "unparseable":
        path.write_text('[users]\nannotator_a = "unclosed\n')
    elif kind == "no_users":
        path.write_text("[server]\nport = 8501\n")

    return path


@pytest.mark.parametrize(
    "kind", ["absent", "directory", "encoding", "unparseable", "no_users"]
)
def test_download_form_withholds_the_export_on_unreadable_secrets(tmp_path, kind):
    path = unreadable(tmp_path, kind)

    app = submit(run_form(path), "annotator_a", "sept-chevaux")
    message = error_body(app)

    assert not app.exception
    assert not app.download_button
    assert message == MESSAGE_UNAVAILABLE
    assert str(path) not in message
    assert "annotator_a" not in message


def test_download_form_withholds_the_export_when_the_workbook_cannot_be_built(
    secrets, caplog
):
    illegal = pd.DataFrame({"n": [1], "note_comment_a": ["patient \x0b Mme Dupont"]})

    app = submit(run_form(secrets, data=illegal), "annotator_a", USERS["annotator_a"])
    message = error_body(app)

    assert not app.exception
    assert not app.download_button
    assert message == MESSAGE_UNAVAILABLE
    assert "Dupont" not in message
    assert "Dupont" not in caplog.text
    assert "note_comment_a" in caplog.text


def test_download_form_builds_no_workbook_before_the_pair_matches(secrets, monkeypatch):
    built = []

    def spy(data, sheet_name):
        built.append(sheet_name)
        return BytesIO(b"xlsx")

    monkeypatch.setattr("edscrib.export.build_workbook", spy)

    run_form(secrets)

    assert not built

    submit(run_form(secrets), "annotator_a", "trois-oranges")

    assert not built

    app = submit(run_form(secrets), "annotator_a", USERS["annotator_a"])

    assert app.download_button
    assert built == [SHEET]
