"""Shaping an annotation frame into an Excel workbook, and gating its download."""

import logging
from io import BytesIO
from pathlib import Path
from typing import Literal

import pandas as pd
import streamlit as st
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE, Cell
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from edscrib.auth import (
    LABEL_PASSWORD,
    LABEL_USERNAME,
    MESSAGE_REJECTED,
    MESSAGE_UNAVAILABLE,
    SecretsError,
    verify_credentials,
)

_logger = logging.getLogger(__name__)

_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_MAX_COLUMN_WIDTH = 60

_MAX_SHEET_NAME = 31

_MAX_CELL_LENGTH = 32767


def _finish_sheet(sheet: Worksheet) -> None:
    """Type, centre, border and widen every cell, then filter on the header row.

    A cell whose text opens on an equals sign is written as a formula, which is what
    `openpyxl` makes of any such string, and it is demoted back to text here. One equal
    to an Excel error code is written as an error cell, by a second branch of the same
    binding, and is demoted with it. An annotation comment reading `=> voir le
    compte-rendu` otherwise reaches the operator as `#NAME?` and reads back as nothing
    at all, silently, on the one field the annotation exists to collect, and one reading
    `#N/A` reaches it as an error for the same silence. The demotion holds because a
    text-typed cell is written without a formula element, and it is what decides
    evaluation rather than the literal content.

    `#N/A` still reads back as nothing through `pandas` once demoted, its default
    `na_values` carrying the literal, as it carries `NA`, `N/A`, `null` and `None` in
    any text cell whatever its type. The cell is then right on disk and right in the
    spreadsheet, and a reader reconciling the export against its source passes
    `keep_default_na=False`. That is the consumer's to pass, no cell typing reaching it.

    It happens in this pass rather than its own: the traversal is the expensive half of
    the export, and there is no reason to pay for it twice. The header is frozen for the
    same traversal's sake: the filter posed on the row above is read from a control that
    scrolls away with it on the first page, which is where a reviewer starts wanting it.

    What the traversal is typed to yield is not only cells, `iter_rows` giving
    `Cell | MergedCell`, and a merged position declares no `data_type` among its slots,
    so the assignment does not type without the narrowing. No merged position reaches
    it: a multi-level header is the one thing that produces one, and `pandas` refuses to
    write it beside the `index=False` this export is written with. The narrowing answers
    the checker, and the branch it guards is dead at runtime, the expected cost of a
    contract carrying a static form and a runtime one.

    A width is capped, the columns being widened holding free text no length bounds.
    The schema puts no ceiling on the attribute, but Office requires it between 0 and
    255 (MS-OI29500 18.3.1.13), which a comment of some two hundred and thirty
    characters already exceeds, and nothing here validates what it writes: the file
    then leaves the deployment carrying a width outside what its reader accepts.
    The cap sits well under that, a column as wide as Office allows taking a screen of
    its own and leaving a reviewer nothing to scan; what it clips stays in the cell,
    and a reader widens the column back or reads the value where it is entered.
    """
    align = Alignment(horizontal="center", vertical="center")

    side = Side(style="thin", color="A5A5A5")

    border = Border(left=side, right=side, top=side, bottom=side)

    fill = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")

    for index, cells in enumerate(sheet.columns, start=1):
        length = max(len(str(cell.value)) for cell in cells)
        width = min((length + 2) * 1.1, _MAX_COLUMN_WIDTH)
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row in sheet.iter_rows(
        min_row=1,
        max_col=sheet.max_column,
        max_row=sheet.max_row,
    ):
        for cell in row:
            if isinstance(cell, Cell) and cell.data_type in {"f", "e"}:
                cell.data_type = "s"

            cell.alignment = align
            cell.border = border

    for cell in sheet[1]:
        cell.fill = fill

    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"

    sheet.freeze_panes = "A2"


def build_workbook(data: pd.DataFrame, sheet_name: str) -> BytesIO:
    """Write a frame into an in-memory xlsx, formatted, and rewind the buffer.

    The index is not written: what an annotation output carries in its columns is the
    export, and a row position is not one of them.

    The buffer is rewound because the caller hands it to a reader that starts where it
    is left, and a workbook read from its end is an empty download.

    The sheet is taken by creation order rather than by the name asked for, which is
    not the name it necessarily carries: the writer titles it before renaming it, and
    `openpyxl` renames a title colliding case-insensitively with one already in the
    book, its own default title included. Asking for `sheet` therefore yields `sheet1`
    and a lookup by name that finds nothing. One frame is written into a writer that
    holds no other sheet, so the last is the one just written.

    Three limits are checked here rather than left to the writing, which passes the
    first two: a title over thirty-one characters `openpyxl` keeps, against a workbook
    the spreadsheet may then refuse, and a value over the cell limit it cuts, against a
    comment that leaves the deployment shorter than the one the annotator wrote. The
    cut is announced, by `pandas` and not by `openpyxl`, in a `UserWarning` nobody is
    listening for. All three are refusals rather than repairs, the export being a copy
    of a gold standard and a copy that quietly differs being worse than one that does
    not arrive.

    The third is a control character, which `openpyxl` does refuse, and refuses by
    naming the whole cell: caught at the display boundary, that message carries the
    clinical text of a named patient into the server log, whose retention and readers
    are those of the deployment's monitoring rather than those of the annotation. It is
    refused here instead, naming the column and the position and never the value, which
    is what locating the cell takes. A column name is read for the same character,
    being written as a header and refused on the same grounds.

    Text is what all three walk, and `kind` decides it: `O` is where a string lands
    under this package's own reader, `str` and `category` included, and `U` is where an
    arrow-backed one does, which a consumer passing `dtype_backend="pyarrow"` hands in.
    Measuring `O` alone leaves that consumer's oversized comment to be cut in silence.
    No number or timestamp reaches either limit once rendered.
    """
    if len(sheet_name) > _MAX_SHEET_NAME:
        raise ValueError(
            f"A sheet name takes at most {_MAX_SHEET_NAME} characters, "
            f"and this one takes {len(sheet_name)} characters"
        )

    for name in data.columns:
        if ILLEGAL_CHARACTERS_RE.search(str(name)):
            raise ValueError(
                f"A column name carries a control character no worksheet accepts, "
                f"and {name!r} carries one"
            )

    for name, column in data.items():
        if column.dtype.kind not in {"O", "U"}:
            continue

        text = column.astype("string")

        longest = text.str.len().max()

        if pd.notna(longest) and longest > _MAX_CELL_LENGTH:
            raise ValueError(
                f"A cell takes at most {_MAX_CELL_LENGTH} characters, "
                f"and column {name} holds one of {longest} characters"
            )

        carries = text.str.contains(ILLEGAL_CHARACTERS_RE, regex=True, na=False)

        if carries.any():
            raise ValueError(
                f"A cell carries a control character no worksheet accepts, "
                f"and column {name} carries one at position "
                f"{int(carries.to_numpy().argmax())}"
            )

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        data.to_excel(writer, index=False, sheet_name=sheet_name)
        _finish_sheet(writer.book.worksheets[-1])

    buffer.seek(0)

    return buffer


def download_form(
    path: str | Path,
    data: pd.DataFrame,
    *,
    filename: str,
    label: str,
    info: str = "",
    icon: str = ":material/download:",
) -> None:
    """Ask for a pair, and serve the export once it matches the deployment secrets.

    This is the body of the dialog `download` opens, and it sits at module level
    because `AppTest` does not run the body of an `st.dialog` (streamlit#9786): held
    inside the decorated shell, none of it would be reachable by a test.

    The pair is checked against the same secrets file as the login gate, so taking a
    gold standard out of the deployment is a deliberate second gate rather than one
    click away in a tab left open. It is the same `[users]` table and no other, so what
    it asks for a second time is authentication and never authorisation: every annotator
    who reaches the app reaches the whole gold standard, for every patient in it. What
    the gate buys is that a tab left open on an unattended machine does not, and a
    deployment wanting the export held to some annotators needs a distinction this file
    does not carry. Which is also its cadence: the file is read on every
    rerun for as long as a password is typed, so a file that cannot be read withholds
    the button and never raises. An operator editing it to add a colleague would
    otherwise paint a traceback over the dialog of an annotator mid-session.

    What the gate holds back is the link and not the bytes. Rendering the button puts
    the workbook in Streamlit's media store and serves it at a URL its endpoint answers
    with no session, cookie or credential of any kind: the path is a 128-bit digest of
    the content, so nobody reaches it by guessing, and it stops answering once no
    session renders the button any more. Until then it is whoever holds it, which is
    the browser history of the annotator and any proxy the deployment sits behind. The
    package cannot close this, Streamlit exposing no other way to serve a download, so
    the operator reads it here rather than assuming the gate covers more than it does.

    That digest is taken over the bytes, and the bytes carry the moment the workbook
    was written, so two builds of one frame hash apart. What the store holds is one
    entry per rerun that reaches the button and not one per export, each at a live URL
    of its own. They stop answering together when the dialog closes, so the exposure
    lasts the dialog and no longer; its width is the number of reruns.

    Neither input carries a `key`, so its value leaves the session when the dialog
    closes and the next opening asks again. `persist_state` is passed rather than
    inherited: the value is a password in clear, so an upstream change of the default
    has to fail a test rather than quietly keep it in server memory.

    The workbook is built once the pair matches and not before: it formats every cell
    of the frame, and each rerun above would otherwise pay for it, the ones where
    nobody has authenticated included. The button does not rerun on its click for the
    same reason: the download takes the file the render already registered, so the
    rerun a click fires by default reaches this line again and pays a second time for
    a workbook nobody collects. It is left eager rather than deferred to the click,
    which would build once but hand the failure to Streamlit's own handler, and with
    it an English message and a log line the operator does not read.

    That build is the second boundary of this shell, and it catches broadly for what it
    holds rather than for what it expects. Uncaught, `client.showErrorDetails` paints a
    message and the server's paths into the browser, and what the build raises is not
    enumerable: `build_workbook`'s own refusals, and under them a writing that refuses a
    timezone-aware datetime, a multi-level header and a frame past the sheet's own
    dimensions, each in its own type. The detail goes to the module logger and one
    message is rendered, against the one the secrets failure already renders: neither is
    anything the annotator can act on. Streamlit's own control flow descends from
    `BaseException` and passes through.

    What reaches that logger is what `build_workbook` names, which is a column and a
    position and never a value. The control character is refused there for that reason:
    left to the writing, it comes back naming the whole cell, and the clinical text of a
    named patient lands in the deployment's log.
    """
    username = st.text_input(label=LABEL_USERNAME, persist_state=None)

    password = st.text_input(label=LABEL_PASSWORD, type="password", persist_state=None)

    if not password:
        return

    try:
        granted = verify_credentials(path, username or "", password)
    except SecretsError:
        _logger.exception("Secrets unusable while checking an export request")
        st.error(MESSAGE_UNAVAILABLE)
        return

    if not granted:
        st.error(MESSAGE_REJECTED)
        return

    if info:
        st.markdown(info)

    st.space("stretch")

    try:
        st.download_button(
            label=label,
            type="primary",
            width="stretch",
            data=build_workbook(data, filename),
            file_name=f"{filename}.xlsx",
            mime=_MIME,
            icon=icon,
            on_click="ignore",
        )
    except Exception:
        _logger.exception("The export could not be built")
        st.error(MESSAGE_UNAVAILABLE)


def download(
    path: str | Path,
    data: pd.DataFrame,
    *,
    filename: str,
    label: str,
    info: str = "",
    icon: str = ":material/download:",
    button_type: Literal["primary", "secondary", "tertiary"] = "secondary",
    dialog_title: str = " ",
    key_button: str = "button-export",
) -> None:
    """Render the button that opens the export dialog.

    `filename` names both the worksheet and the file the browser receives.

    `dialog_title` holds a single space because `st.dialog` takes a title and the form
    needs none: what it asks for is legible from its own two fields.

    The button carries a key so that a page offering the export twice, in a sidebar and
    again in its body, names which of the two to move rather than failing on an
    identifier Streamlit derived from parameters the two calls share. The form the
    dialog opens takes none: only one dialog is ever open, and the two fields it holds
    are keyless on purpose, that being what takes the password out of the session when
    it closes.

    What the export carries is the frame of the run the button was clicked in, not of
    the moment it is downloaded. `st.dialog` is a fragment, and a fragment rerun calls
    the body Streamlit stored rather than this function again, so every rerun the open
    dialog drives re-enters a closure holding that frame. Nothing the annotator does
    can move it, the dialog being modal for them, but the output is shared: a colleague
    saving meanwhile is not in what leaves. Reopening the dialog is what takes it.
    """

    @st.dialog(title=dialog_title, width="medium")
    def export_dialog() -> None:
        download_form(path, data, filename=filename, label=label, info=info, icon=icon)

    if st.button(
        label=label, type=button_type, width="stretch", icon=icon, key=key_button
    ):
        export_dialog()
